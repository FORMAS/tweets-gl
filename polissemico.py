from openpyxl import Workbook
import openpyxl
import sys

book = openpyxl.load_workbook('tfreq-gl-filter.xlsx')

sheet = book.active
i = 0
j = 0

termos = []
polissemicos = []
for r in sheet.iter_rows():
    # Ignora a primeira linha (label)
    if i != 0:
        #print(r[1].value)
        termos.append(r[1].value)
    i += 1

# COMPARE
for t in termos:
    for p in termos:
        if t == p:
            #print(t)
            j += 1
    if j > 1:
        polissemicos.append(t)
    j = 0

i = 1
print(len(polissemicos))
exit()
print('letsgo')
for r in sheet.iter_rows():
    # Ignora a primeira linha (label)
    if i != 0:
        #print(r[1].value)
        if r[1].value in polissemicos:
            sheet['I' + str(i)] = 'SIM'
            polissemicos.remove(r[1].value)
        else:
            sheet['I' + str(i)] = 'não'
        '''for p in polissemicos:
            if r[1].value == p:
                #if r[8].value != 'SIM':
                print(i, p)
                #i += 1
                #exit()
                sheet['I' + str(i)] = 'SIM'
                #book.save('poli-tfreq-gl-filter.xlsx')
                i+=1
                polissemicos.remove(p)
                #print('remove ', p)
                #exit()
        '''
    i += 1
book.save('poli-tfreq-gl-filter.xlsx')

#print(polissemicos)
#print(len(polissemicos))